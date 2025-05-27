import os
import pandas as pd
import openpyxl
import json
from typing import Dict, List, Any, Optional, Union

class ExcelLoader:
    """Tool for loading and extracting content from Excel files"""
    
    def __init__(self, file_path: str):
        """Initialize with the path to an Excel file"""
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"Excel file not found: {file_path}")
        
        self.file_path = file_path
        self.workbook = openpyxl.load_workbook(file_path, data_only=True)
    
    def get_metadata(self) -> Dict[str, Any]:
        """Extract metadata from the Excel file"""
        return {
            "sheet_names": self.workbook.sheetnames,
            "sheet_count": len(self.workbook.sheetnames),
            "file_size": os.path.getsize(self.file_path),
            "properties": {
                "creator": self.workbook.properties.creator,
                "last_modified_by": self.workbook.properties.lastModifiedBy,
                "created": str(self.workbook.properties.created) if self.workbook.properties.created else None,
                "modified": str(self.workbook.properties.modified) if self.workbook.properties.modified else None,
            }
        }
    
    def list_sheets(self) -> List[str]:
        """List all sheet names in the Excel file"""
        return self.workbook.sheetnames
    
    def extract_sheet_data(self, sheet_name: Optional[str] = None, max_rows: Optional[int] = None) -> Dict[str, Any]:
        """
        Extract data from a specific sheet as a list of dictionaries
        
        Args:
            sheet_name: Name of the sheet to extract. If None, uses the active sheet.
            max_rows: Maximum number of rows to extract. If None, extracts all rows.
        
        Returns:
            Dictionary with sheet data and metadata
        """
        # Use the specified sheet or the active sheet
        if sheet_name is None:
            sheet = self.workbook.active
        else:
            if sheet_name not in self.workbook.sheetnames:
                raise ValueError(f"Sheet '{sheet_name}' not found in workbook")
            sheet = self.workbook[sheet_name]
        
        # Get headers from the first row
        headers = []
        for cell in sheet[1]:
            headers.append(str(cell.value) if cell.value is not None else f"Column_{cell.column_letter}")
        
        # Extract data
        data = []
        for i, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if max_rows is not None and i > max_rows + 1:  # +1 because we start from row 2
                break
                
            row_dict = {}
            for j, value in enumerate(row):
                if j < len(headers):
                    row_dict[headers[j]] = value
                else:
                    row_dict[f"Column_{j+1}"] = value
            data.append(row_dict)
        
        return {
            "sheet_name": sheet.title,
            "headers": headers,
            "row_count": len(data),
            "data": data
        }
    
    def extract_sheet_as_df(self, sheet_name: Optional[str] = None) -> pd.DataFrame:
        """
        Extract data from a specific sheet as a pandas DataFrame
        
        Args:
            sheet_name: Name of the sheet to extract. If None, uses the active sheet.
        
        Returns:
            Pandas DataFrame with the sheet data
        """
        return pd.read_excel(self.file_path, sheet_name=sheet_name)
    
    def query_data(self, sheet_name: Optional[str] = None, query: str = "") -> Dict[str, Any]:
        """
        Query data using pandas operations
        
        Args:
            sheet_name: Name of the sheet to query. If None, uses the active sheet.
            query: A query string for pandas eval (e.g., "data[data['Column'] > 100]")
        
        Returns:
            Result of the query operation
        """
        df = self.extract_sheet_as_df(sheet_name)
        
        # Safety check - only allow certain operations
        if any(op in query.lower() for op in ["delete", "drop", "exec", "eval"]):
            raise ValueError("Potentially unsafe operation detected in query")
        
        # Create a local context with the DataFrame
        local_context = {"data": df}
        
        # Execute the query
        try:
            result = eval(query, {"__builtins__": {}}, local_context)
            
            # Convert result to serializable format
            if isinstance(result, pd.DataFrame):
                return {
                    "type": "dataframe",
                    "columns": result.columns.tolist(),
                    "data": result.to_dict(orient="records"),
                    "shape": result.shape
                }
            elif isinstance(result, pd.Series):
                return {
                    "type": "series",
                    "name": result.name,
                    "data": result.to_dict(),
                    "shape": result.shape
                }
            else:
                # Try to convert to JSON-serializable format
                return {
                    "type": str(type(result).__name__),
                    "result": json.loads(json.dumps(result, default=str))
                }
        except Exception as e:
            return {"error": str(e)}
    
    def close(self) -> None:
        """Close the Excel workbook"""
        if hasattr(self, 'workbook') and self.workbook:
            self.workbook.close()
    
    def __del__(self):
        """Ensure workbook is closed on deletion"""
        self.close()

# Usage as a LangChain tool
def create_excel_loader_tools(file_path):
    """Create a set of Excel tools for use with LangChain"""
    from langchain.tools import Tool
    
    excel_loader = ExcelLoader(file_path)
    
    tools = [
        Tool(
            name="list_excel_sheets",
            func=lambda _: excel_loader.list_sheets(),
            description="List all sheets in the Excel file."
        ),
        Tool(
            name="get_excel_metadata",
            func=lambda _: excel_loader.get_metadata(),
            description="Get metadata about the Excel file."
        ),
        Tool(
            name="extract_excel_sheet",
            func=lambda args: excel_loader.extract_sheet_data(
                sheet_name=args.get("sheet_name"),
                max_rows=args.get("max_rows")
            ),
            description="Extract data from a specific Excel sheet. Input should be a dict with 'sheet_name' (optional) and 'max_rows' (optional)."
        ),
        Tool(
            name="query_excel_data",
            func=lambda args: excel_loader.query_data(
                sheet_name=args.get("sheet_name"),
                query=args.get("query", "")
            ),
            description="Query Excel data using pandas operations. Input should be a dict with 'sheet_name' (optional) and 'query' (required)."
        )
    ]
    
    return tools